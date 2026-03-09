from flask import Flask, render_template, request, jsonify, send_file
from llm_parser import parse_plan
from normalizer import normalize
from rule_engine import calculate_total, PHASE_RATIOS
from config_loader import load_config
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.chart.series import DataPoint
import io
import json
import os
from datetime import datetime

app = Flask(__name__)
CONFIG_PATH = os.path.join(os.path.dirname(__file__), "config_params.json")


# ─── 样式常量 ────────────────────────────────────────────────

C_BLUE_DARK   = "1A3C6E"   # 深蓝  - 主标题
C_BLUE        = "1a56db"   # 品牌蓝 - 表头
C_BLUE_LIGHT  = "D6E4FF"   # 浅蓝  - 交替行
C_ORANGE      = "E67E22"   # 橙色  - 重点数字
C_GREEN       = "1E8449"   # 绿色  - 小计/合计
C_GRAY        = "F5F7FA"   # 浅灰  - 背景
C_WHITE       = "FFFFFF"
C_TEXT        = "1A1A2E"
C_GOLD        = "F39C12"   # 金色  - 特别提示


def _side(color="CCCCCC", style="thin"):
    return Side(style=style, color=color)


def _border(color="CCCCCC"):
    s = _side(color)
    return Border(left=s, right=s, top=s, bottom=s)


def _header_border():
    s = _side("FFFFFF", "medium")
    return Border(left=s, right=s, top=s, bottom=s)


def _font(bold=False, size=11, color=C_TEXT, name="微软雅黑"):
    return Font(bold=bold, size=size, color=color, name=name)


def _fill(color):
    return PatternFill("solid", fgColor=color)


def _center(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)


def _left(wrap=False):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)


def _right():
    return Alignment(horizontal="right", vertical="center")


def _num(val, decimals=2):
    return round(float(val), decimals)


def _merge_style(ws, cell_range, value="", bold=False, size=11,
                 color=C_TEXT, bg=None, align="center", wrap=False):
    ws.merge_cells(cell_range)
    top_left = ws[cell_range.split(":")[0]]
    top_left.value = value
    top_left.font = _font(bold=bold, size=size, color=color)
    if bg:
        top_left.fill = _fill(bg)
    a = _center(wrap) if align == "center" else _left(wrap) if align == "left" else _right()
    top_left.alignment = a
    return top_left


def _row_bg(ws, row, start_col, end_col, color):
    for c in range(start_col, end_col + 1):
        ws.cell(row, c).fill = _fill(color)


def _apply_header_row(ws, row_idx, headers, fills=None):
    """渲染表头行，fills 为每列背景色列表（可选）"""
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=h)
        cell.font = _font(bold=True, size=10, color=C_WHITE)
        cell.fill = _fill(fills[col_idx - 1] if fills else C_BLUE)
        cell.alignment = _center(wrap=True)
        cell.border = _header_border()


def _apply_data_row(ws, row_idx, values, alt=False, bold=False, color=None):
    bg = C_BLUE_LIGHT if alt else C_WHITE
    for col_idx, v in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=v)
        cell.font = _font(bold=bold, size=10, color=color or C_TEXT)
        cell.fill = _fill(bg)
        cell.alignment = _center() if isinstance(v, (int, float)) else _left()
        cell.border = _border()


# ─── 路由 ────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/config", methods=["GET"])
def get_config():
    return jsonify(load_config())


@app.route("/config", methods=["POST"])
def save_config():
    data = request.json
    with open(CONFIG_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    return jsonify({"ok": True})


@app.route("/parse", methods=["POST"])
def parse():
    data = request.json
    result = parse_plan(data.get("md", ""))
    return jsonify({"result": result})


@app.route("/batch_evaluate", methods=["POST"])
def batch_evaluate():
    req = request.json
    files = req.get("files", [])
    config = load_config()
    results = []

    for f in files:
        name = f.get("name", "未命名")
        content = f.get("content", "")
        try:
            parsed_str = parse_plan(content)
            cleaned = parsed_str.strip()
            if cleaned.startswith("```"):
                lines = cleaned.split("\n")
                cleaned = "\n".join(lines[1:])
            if cleaned.endswith("```"):
                cleaned = cleaned.rsplit("```", 1)[0]
            parsed = json.loads(cleaned)
            normalized = normalize(parsed)
            evaluated = calculate_total(normalized, config)
            results.append({"name": name, "status": "ok", "data": evaluated})
        except Exception as e:
            results.append({"name": name, "status": "error", "error": str(e)})

    return jsonify({"results": results})


@app.route("/export_excel", methods=["POST"])
def export_excel():
    data = request.json
    file_stream = generate_excel(data)
    return send_file(
        file_stream,
        as_attachment=True,
        download_name="信创适配费用评估报告.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# ─── Excel 生成 ───────────────────────────────────────────────

def generate_excel(batch_results):
    wb = Workbook()
    ok_results = [r for r in batch_results if r.get("status") == "ok"]

    _sheet_cover(wb, ok_results)
    _sheet_summary(wb, ok_results)
    _sheet_cost_breakdown(wb, ok_results)
    _sheet_phase(wb, ok_results)
    for r in ok_results:
        _sheet_system_detail(wb, r)
    _sheet_params(wb)

    # 删除默认 Sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    return file_stream


def _sheet_cover(wb, ok_results):
    """封面页"""
    ws = wb.create_sheet("封面", 0)
    ws.sheet_view.showGridLines = False

    # 行高列宽
    for r in range(1, 40):
        ws.row_dimensions[r].height = 22
    for c in range(1, 12):
        ws.column_dimensions[get_column_letter(c)].width = 14

    # 顶部色块
    for r in range(1, 8):
        _row_bg(ws, r, 1, 11, C_BLUE_DARK)

    _merge_style(ws, "B2:J3", "国产化适配改造费用评估报告",
                 bold=True, size=22, color=C_WHITE, align="center")
    _merge_style(ws, "B4:J4", "Localization Adaptation Cost Assessment Report",
                 bold=False, size=12, color="A8C4E8", align="center")

    # 分隔线
    for r in range(8, 10):
        _row_bg(ws, r, 1, 11, C_GRAY)

    # 基本信息区
    total_projects = len(ok_results)
    total_cost = sum(r["data"]["cost"]["total_incl_tax"] for r in ok_results)
    total_work = sum(r["data"]["total_workload"] for r in ok_results)
    total_systems = sum(len(r["data"]["systems"]) for r in ok_results)

    info_items = [
        ("评估项目数", str(total_projects)),
        ("涉及子系统数", str(total_systems)),
        ("合计工作量（人天）", f"{round(total_work, 1)}"),
        ("合计费用（含税）", f"¥ {total_cost:,.2f}"),
        ("报告生成日期", datetime.now().strftime("%Y年%m月%d日")),
        ("评估依据", "GBT 16680、信创适配指导规范"),
    ]

    for i, (label, val) in enumerate(info_items):
        row = 11 + i * 2
        _merge_style(ws, f"B{row}:D{row}", label, bold=True, size=11,
                     color=C_BLUE_DARK, align="left")
        _merge_style(ws, f"E{row}:J{row}", val, bold=False, size=11,
                     color=C_TEXT, align="left")

    # 技术栈对照（汇总所有项目）
    all_existing = {}
    all_target = {}
    for r in ok_results:
        d = r["data"]
        for k, v in d.get("existing_tech_stack", {}).items():
            if v and k not in all_existing:
                all_existing[k] = v
        for k, v in d.get("target_tech_stack", {}).items():
            if v and k not in all_target:
                all_target[k] = v

    if all_existing or all_target:
        tech_start = 24
        _merge_style(ws, f"B{tech_start}:J{tech_start}",
                     "技术栈改造对照",
                     bold=True, size=11, color=C_WHITE, bg=C_BLUE, align="center")
        tech_start += 1
        _merge_style(ws, f"B{tech_start}:E{tech_start}", "改造维度",
                     bold=True, size=10, color=C_WHITE, bg=C_BLUE_DARK, align="center")
        _merge_style(ws, f"F{tech_start}:G{tech_start}", "现有技术栈",
                     bold=True, size=10, color=C_WHITE, bg=C_BLUE_DARK, align="center")
        _merge_style(ws, f"H{tech_start}:J{tech_start}", "改造目标（信创）",
                     bold=True, size=10, color=C_WHITE, bg=C_BLUE_DARK, align="center")
        tech_start += 1
        label_map = {"database": "数据库", "os": "操作系统", "middleware": "中间件",
                     "hardware": "硬件平台", "frontend": "前端框架", "etl": "ETL平台"}
        all_keys = list(dict.fromkeys(list(all_existing.keys()) + list(all_target.keys())))
        for ki, key in enumerate(all_keys):
            bg = C_BLUE_LIGHT if ki % 2 == 0 else C_WHITE
            label = label_map.get(key, key)
            _merge_style(ws, f"B{tech_start}:E{tech_start}", label, bg=bg, align="left")
            _merge_style(ws, f"F{tech_start}:G{tech_start}", all_existing.get(key, "—"), bg=bg, align="left")
            _merge_style(ws, f"H{tech_start}:J{tech_start}", all_target.get(key, "—"), bg=bg, color="1E8449", align="left")
            tech_start += 1

    # 免责声明
    note_row = max(tech_start + 1, 34)
    _merge_style(ws, f"B{note_row}:J{note_row + 2}",
                 "【说明】本报告依据建设方案文档，结合信创改造工作量模型自动生成，"
                 "仅作为项目预算参考依据，最终费用以正式合同及详细方案为准。",
                 bold=False, size=9, color="888888", bg=C_GRAY, align="left", wrap=True)


def _sheet_summary(wb, ok_results):
    """汇总表（含费用构成说明与占比）"""
    ws = wb.create_sheet("费用汇总")
    ws.sheet_view.showGridLines = False

    col_widths = [5, 26, 8, 10, 14, 14, 14, 12, 12, 14, 14, 8]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 14
    ws.row_dimensions[3].height = 36
    ws.row_dimensions[4].height = 36

    TOTAL_COLS = len(col_widths)
    last_col = get_column_letter(TOTAL_COLS)

    # 大标题
    _merge_style(ws, f"A1:{last_col}1", "国产化适配改造费用评估汇总表",
                 bold=True, size=14, color=C_WHITE, bg=C_BLUE_DARK, align="center")
    _merge_style(ws, f"A2:{last_col}2",
                 f"报告生成时间：{datetime.now().strftime('%Y年%m月%d日 %H:%M')}　　评估依据：模块复杂度+技术改造附加量模型（符合GBT 16680规范）",
                 bold=False, size=9, color="666666", bg=C_GRAY, align="left")

    headers = [
        "序号", "项目 / 文件名称", "子系统\n数量", "总工作量\n(人天)",
        "人工费\n(元)", "测试费\n(元)", "人工+测试\n小计(元)",
        "项目管理费\n(元)", "风险费\n(元)",
        "不含税合计\n(元)", "含税合计\n(元)", "占比"
    ]
    _apply_header_row(ws, 3, headers)

    grand = {k: 0.0 for k in ["total_workload", "labor_cost", "test_cost",
                                "subtotal", "management_cost", "risk_cost",
                                "total_excl_tax", "total_incl_tax"]}
    for r in ok_results:
        d = r["data"]
        c = d["cost"]
        grand["total_workload"] += d.get("total_workload", 0)
        for k in ["labor_cost", "test_cost", "subtotal", "management_cost", "risk_cost", "total_excl_tax", "total_incl_tax"]:
            grand[k] += c.get(k, 0)

    for idx, r in enumerate(ok_results, 1):
        d = r["data"]
        c = d["cost"]
        alt = (idx % 2 == 0)
        pct = (c["total_incl_tax"] / grand["total_incl_tax"] * 100) if grand["total_incl_tax"] else 0
        row = [
            idx, r["name"].replace(".md", ""), len(d["systems"]),
            _num(d["total_workload"]),
            _num(c["labor_cost"]), _num(c["test_cost"]), _num(c["subtotal"]),
            _num(c["management_cost"]), _num(c["risk_cost"]),
            _num(c["total_excl_tax"]), _num(c["total_incl_tax"]),
            f"{pct:.1f}%"
        ]
        _apply_data_row(ws, 4 + idx - 1, row, alt=alt)

    # 合计行
    total_row = 4 + len(ok_results)
    for col_idx in range(1, TOTAL_COLS + 1):
        cell = ws.cell(row=total_row, column=col_idx)
        cell.font = _font(bold=True, size=11, color=C_WHITE)
        cell.fill = _fill(C_BLUE_DARK)
        cell.alignment = _center()
        cell.border = _header_border()
    total_values = [
        "", "合  计", "", _num(grand["total_workload"]),
        _num(grand["labor_cost"]), _num(grand["test_cost"]), _num(grand["subtotal"]),
        _num(grand["management_cost"]), _num(grand["risk_cost"]),
        _num(grand["total_excl_tax"]), _num(grand["total_incl_tax"]), "100%"
    ]
    for col_idx, v in enumerate(total_values, 1):
        ws.cell(row=total_row, column=col_idx).value = v

    # 大写金额
    note_row = total_row + 1
    _merge_style(ws, f"A{note_row}:{last_col}{note_row}",
                 f"含税合计（大写）：{_amount_cn(grand['total_incl_tax'])}",
                 bold=True, size=11, color=C_ORANGE, bg="FFF8E7", align="left")
    ws.row_dimensions[note_row].height = 20

    # ── 费用构成说明区 ──
    sec_row = note_row + 2
    _merge_style(ws, f"A{sec_row}:{last_col}{sec_row}",
                 "费用构成模型说明",
                 bold=True, size=11, color=C_WHITE, bg=C_BLUE, align="center")
    ws.row_dimensions[sec_row].height = 20
    sec_row += 1

    config = load_config()
    price = config.get("price_per_day", 2200)
    test_r = config.get("test_rate", 0.15)
    mgmt_r = config.get("management_rate", 0.12)
    risk_r = config.get("risk_rate", 0.10)
    tax_r = config.get("tax_rate", 0.09)

    model_lines = [
        ("工作量公式",
         "系统工作量 = (模块数×系数 + 接口数×系数 + 数据表数×系数 + 数据量×系数 + 技术改造附加量) × 级别系数 × 复杂度系数"),
        ("人工费",
         f"人工费 = 总工作量（人天） × 日综合费率（¥{price:,}/天），含工资、社保、管理分摊"),
        ("测试费",
         f"测试费 = 人工费 × {test_r*100:.0f}%，含功能测试、兼容性测试、回归测试"),
        ("项目管理费",
         f"项目管理费 = (人工费+测试费) × {mgmt_r*100:.0f}%，含项目经理、进度管控、沟通协调"),
        ("风险不可预见费",
         f"风险费 = (人工费+测试费) × {risk_r*100:.0f}%，用于覆盖技术不确定性及需求变更"),
        ("增值税",
         f"含税价 = 不含税价 × (1 + {tax_r*100:.0f}%)，信息技术服务适用税率"),
        ("级别系数",
         "核心系统 × 1.3 | 重要系统 × 1.1 | 一般系统 × 1.0"),
        ("复杂度系数",
         "基础1.0；模块数>20 +0.15；接口数>30 +0.10 / >10 +0.05；改造类型≥4 +0.20 / ≥2 +0.10（可叠加）"),
    ]

    _apply_header_row(ws, sec_row, ["费用项目", "计算说明"],
                      fills=[C_BLUE_DARK, C_BLUE_DARK])
    ws.merge_cells(f"B{sec_row}:{last_col}{sec_row}")
    sec_row += 1
    for ki, (label, desc) in enumerate(model_lines):
        alt = ki % 2 == 0
        bg = C_BLUE_LIGHT if alt else C_WHITE
        cell_l = ws.cell(sec_row, 1, label)
        cell_l.font = _font(bold=True, size=9, color=C_BLUE_DARK)
        cell_l.fill = _fill(bg)
        cell_l.alignment = _left()
        cell_l.border = _border()
        ws.merge_cells(f"B{sec_row}:{last_col}{sec_row}")
        cell_r = ws.cell(sec_row, 2, desc)
        cell_r.font = _font(size=9, color=C_TEXT)
        cell_r.fill = _fill(bg)
        cell_r.alignment = _left(wrap=True)
        cell_r.border = _border()
        ws.row_dimensions[sec_row].height = 18
        sec_row += 1


def _sheet_cost_breakdown(wb, ok_results):
    """工作量拆解明细表（含各分项贡献及改造类型附加明细）"""
    ws = wb.create_sheet("工作量拆解明细")
    ws.sheet_view.showGridLines = False

    config = load_config()
    adapt_cfg = config.get("adaptation_factors", {})

    col_widths = [4, 22, 6, 8, 8, 8, 8, 10, 10, 10, 8, 8, 10, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    TOTAL_COLS = len(col_widths)
    last_col = get_column_letter(TOTAL_COLS)

    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 14
    ws.row_dimensions[3].height = 42

    _merge_style(ws, f"A1:{last_col}1", "系统工作量拆解明细表",
                 bold=True, size=14, color=C_WHITE, bg=C_BLUE_DARK, align="center")
    _merge_style(ws, f"A2:{last_col}2",
                 "工作量计算公式：(模块工作量 + 接口工作量 + 数据表工作量 + 数据迁移量 + 改造类型附加量) × 级别系数 × 复杂度系数",
                 bold=False, size=9, color="555555", bg=C_GRAY, align="left")

    headers = [
        "序\n号", "系统名称", "级别", "模块\n工作量\n(人天)", "接口\n工作量\n(人天)",
        "数据表\n工作量\n(人天)", "数据迁移\n工作量\n(人天)",
        "基础\n小计\n(人天)", "改造类型\n附加量\n(人天)", "合计前\n小计\n(人天)",
        "级别\n系数", "复杂度\n系数", "最终工作量\n(人天)", "占项目\n工作量"
    ]
    _apply_header_row(ws, 3, headers)

    row_idx = 4
    seq = 1
    for r in ok_results:
        d = r["data"]
        proj_work = d["total_workload"] or 1
        # 项目分组标题
        _merge_style(ws, f"A{row_idx}:{last_col}{row_idx}",
                     f"  项目：{r['name'].replace('.md', '')}",
                     bold=True, size=10, color=C_WHITE, bg=C_BLUE, align="left")
        ws.row_dimensions[row_idx].height = 18
        row_idx += 1

        for sys in d["systems"]:
            alt = (seq % 2 == 0)
            mw = sys.get("module_work", 0)
            iw = sys.get("interface_work", 0)
            dw = sys.get("db_work", 0)
            dtw = sys.get("data_work", 0)
            base = sys.get("base_work", 0)
            adapt = sys.get("adaptation_work", 0)
            pre_factor = round(base + adapt, 2)
            lf = sys.get("level_factor", 1.0)
            cf = sys.get("complexity", 1.0)
            tw = sys.get("total_work", 0)
            pct = (tw / proj_work * 100)
            row = [
                seq, sys["name"], sys.get("level", "—"),
                _num(mw), _num(iw), _num(dw), _num(dtw),
                _num(base), _num(adapt), _num(pre_factor),
                lf, cf, _num(tw), f"{pct:.1f}%"
            ]
            _apply_data_row(ws, row_idx, row, alt=alt)
            row_idx += 1
            seq += 1

    # ── 改造类型附加量明细子表 ──
    row_idx += 1
    _merge_style(ws, f"A{row_idx}:{last_col}{row_idx}",
                 "改造类型附加量参考标准（依据config_params.json当前参数）",
                 bold=True, size=11, color=C_WHITE, bg=C_BLUE, align="center")
    ws.row_dimensions[row_idx].height = 20
    row_idx += 1

    _apply_header_row(ws, row_idx, ["改造类型", "每系统附加工作量(人天)", "典型工作内容说明", "备注"])
    ws.merge_cells(f"C{row_idx}:{last_col}{row_idx}")
    row_idx += 1

    adapt_desc = {
        "数据库替换": ("国产数据库安装配置、SQL语法兼容性改造、驱动替换、存储过程重写、性能调优", "如Oracle→达梦/人大金仓"),
        "ETL平台适配": ("ETL工具国产化替换、数据抽取转换规则迁移、调度任务重配置", "如Kettle→FineDataLink"),
        "数据迁移": ("历史数据清洗、格式转换、校验比对、迁移脚本编写与执行", "含数据质量核查"),
        "硬件架构适配": ("x86→飞腾/鲲鹏编译适配、底层库替换、性能基准测试", "ARM架构兼容性验证"),
        "操作系统适配": ("应用在麒麟/统信OS部署调试、依赖库兼容、系统调用适配", "含运行环境验证"),
        "安全加固": ("国密算法集成、接口鉴权改造、日志审计、漏洞扫描整改", "SM2/SM4等国密标准"),
        "中间件替换": ("消息队列/缓存/服务治理组件国产替换及接口适配", "如RabbitMQ→RocketMQ"),
        "前端框架适配": ("前端UI组件库兼容性修复、浏览器适配、国产化字体/打印适配", "信创浏览器兼容"),
        "接口改造": ("第三方/外部接口协议适配、联调测试、文档更新", "按系统计，外部接口另计"),
    }
    for ki, (k, v) in enumerate(adapt_cfg.items()):
        alt = ki % 2 == 0
        bg = C_BLUE_LIGHT if alt else C_WHITE
        desc, note = adapt_desc.get(k, ("—", ""))
        cell_k = ws.cell(row_idx, 1, k)
        cell_k.font = _font(bold=True, size=9, color=C_BLUE_DARK); cell_k.fill = _fill(bg); cell_k.alignment = _left(); cell_k.border = _border()
        cell_v = ws.cell(row_idx, 2, v)
        cell_v.font = _font(bold=True, size=10, color=C_ORANGE); cell_v.fill = _fill(bg); cell_v.alignment = _center(); cell_v.border = _border()
        ws.merge_cells(f"C{row_idx}:{get_column_letter(TOTAL_COLS - 1)}{row_idx}")
        cell_d = ws.cell(row_idx, 3, desc)
        cell_d.font = _font(size=9, color=C_TEXT); cell_d.fill = _fill(bg); cell_d.alignment = _left(wrap=True); cell_d.border = _border()
        cell_n = ws.cell(row_idx, TOTAL_COLS, note)
        cell_n.font = _font(size=9, color="888888"); cell_n.fill = _fill(bg); cell_n.alignment = _left(wrap=True); cell_n.border = _border()
        ws.row_dimensions[row_idx].height = 18
        row_idx += 1

    # 说明脚注
    row_idx += 1
    foot_notes = [
        "【级别系数】核心系统×1.3（直接影响业务连续性），重要系统×1.1（关键业务支撑），一般系统×1.0",
        "【复杂度系数】基础1.0；模块数>20 +0.15；接口数>30 +0.10、>10 +0.05；改造类型≥4 +0.20、≥2 +0.10（各项可叠加）",
        "【外部接口专项】项目级外部对接接口工作量 = 外部接口总数 × 接口改造系数，单独计入项目总工作量",
    ]
    for note in foot_notes:
        _merge_style(ws, f"A{row_idx}:{last_col}{row_idx}", note,
                     bold=False, size=9, color="555555", bg=C_GRAY, align="left", wrap=True)
        ws.row_dimensions[row_idx].height = 18
        row_idx += 1


def _sheet_phase(wb, ok_results):
    """工时阶段分解表"""
    ws = wb.create_sheet("工时阶段分解")
    ws.sheet_view.showGridLines = False

    phases = list(PHASE_RATIOS.keys())
    col_widths = [6, 28] + [14] * len(phases) + [16]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 16

    _merge_style(ws, f"A1:{get_column_letter(len(col_widths))}1",
                 "工时阶段分解表（各阶段工作量单位：人天）",
                 bold=True, size=14, color=C_WHITE, bg=C_BLUE_DARK, align="center")

    headers = ["序号", "系统名称"] + phases + ["合计(人天)"]
    phase_colors = [C_BLUE, "1565C0", "0D47A1", "1976D2", "1565C0"]
    fills_h = [C_BLUE_DARK, C_BLUE_DARK] + phase_colors + [C_GREEN.replace("1E8449", "1B5E20")]
    _apply_header_row(ws, 2, headers, fills=[C_BLUE_DARK] * len(headers))

    row_idx = 3
    seq = 1
    for r in ok_results:
        for sys in r["data"]["systems"]:
            alt = (seq % 2 == 0)
            phase_vals = [_num(sys.get("phases", {}).get(p, 0)) for p in phases]
            row = [seq, sys["name"]] + phase_vals + [_num(sys["total_work"])]
            _apply_data_row(ws, row_idx, row, alt=alt)
            row_idx += 1
            seq += 1

    # 合计行
    _row_bg(ws, row_idx, 1, len(headers), C_GREEN)
    ws.cell(row_idx, 1, "").fill = _fill(C_GREEN)
    ws.cell(row_idx, 2).value = "合 计"
    ws.cell(row_idx, 2).font = _font(bold=True, color=C_WHITE)
    ws.cell(row_idx, 2).fill = _fill(C_GREEN)
    ws.cell(row_idx, 2).alignment = _center()
    for pi, phase in enumerate(phases):
        total_phase = sum(
            sys.get("phases", {}).get(phase, 0)
            for r in ok_results for sys in r["data"]["systems"]
        )
        cell = ws.cell(row_idx, 3 + pi, round(total_phase, 2))
        cell.font = _font(bold=True, color=C_WHITE)
        cell.fill = _fill(C_GREEN)
        cell.alignment = _center()

    total_all = sum(r["data"]["total_workload"] for r in ok_results)
    total_cell = ws.cell(row_idx, len(headers), round(total_all, 2))
    total_cell.font = _font(bold=True, color=C_WHITE)
    total_cell.fill = _fill(C_GREEN)
    total_cell.alignment = _center()

    # 比例说明
    row_idx += 2
    _merge_style(ws, f"A{row_idx}:{get_column_letter(len(headers))}{row_idx}",
                 "阶段比例：调研分析15% / 环境搭建10% / 代码适配45% / 测试验证20% / 上线部署10%",
                 bold=False, size=9, color="555555", bg=C_GRAY, align="left")


def _sheet_system_detail(wb, r):
    """每个项目的完整评估明细 Sheet（含系统分析、费用拆解、工时分布）"""
    sheet_name = r["name"].replace(".md", "")[:26] + "-评估明细"
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False

    d = r["data"]
    project_name = d.get("project_name", r["name"])
    background = d.get("project_background", "")
    config = load_config()
    price_per_day = config.get("price_per_day", 2200)
    test_rate = config.get("test_rate", 0.15)
    management_rate = config.get("management_rate", 0.12)
    risk_rate = config.get("risk_rate", 0.10)
    tax_rate = config.get("tax_rate", 0.09)

    TOTAL_COLS = 17
    last_col = get_column_letter(TOTAL_COLS)
    col_widths = [4, 20, 7, 7, 7, 8, 8, 8, 8, 7, 7, 7, 7, 12, 12, 12, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── 封头信息 ──
    _merge_style(ws, f"A1:{last_col}1",
                 f"项目名称：{project_name}",
                 bold=True, size=13, color=C_WHITE, bg=C_BLUE_DARK, align="left")
    ws.row_dimensions[1].height = 22
    row_idx = 2
    if background:
        _merge_style(ws, f"A{row_idx}:{last_col}{row_idx}",
                     f"项目背景：{background}",
                     bold=False, size=9, color="333333", bg=C_GRAY,
                     align="left", wrap=True)
        ws.row_dimensions[row_idx].height = 36
        row_idx += 1

    # ── 技术栈对照 ──
    existing = d.get("existing_tech_stack", {})
    target = d.get("target_tech_stack", {})
    if existing or target:
        _merge_style(ws, f"A{row_idx}:{last_col}{row_idx}",
                     "技术栈改造对照（现有 → 信创目标）",
                     bold=True, size=10, color=C_WHITE, bg=C_BLUE, align="center")
        ws.row_dimensions[row_idx].height = 18
        row_idx += 1
        label_map = {"database": "数据库", "os": "操作系统", "middleware": "中间件",
                     "hardware": "硬件平台", "frontend": "前端框架", "etl": "ETL平台"}
        all_keys = list(dict.fromkeys(list(existing.keys()) + list(target.keys())))
        for ki, key in enumerate(all_keys):
            bg = C_BLUE_LIGHT if ki % 2 == 0 else C_WHITE
            label = label_map.get(key, key)
            _merge_style(ws, f"A{row_idx}:C{row_idx}", label, bold=True, size=9, bg=bg, align="left")
            _merge_style(ws, f"D{row_idx}:I{row_idx}", existing.get(key, "—"), size=9, color="CC3300", bg=bg, align="left")
            _merge_style(ws, f"J{row_idx}:{last_col}{row_idx}", target.get(key, "—"), size=9, color=C_GREEN, bg=bg, align="left")
            ws.row_dimensions[row_idx].height = 16
            row_idx += 1

    # ── 改造难点说明 ──
    has_notes = any(sys.get("complexity_note") for sys in d["systems"])
    if has_notes:
        _merge_style(ws, f"A{row_idx}:{last_col}{row_idx}",
                     "各系统改造难点分析",
                     bold=True, size=10, color=C_WHITE, bg=C_BLUE, align="center")
        ws.row_dimensions[row_idx].height = 18
        row_idx += 1
        for si, sys in enumerate(d["systems"], 1):
            note = sys.get("complexity_note", "")
            if note:
                _merge_style(ws, f"A{row_idx}:{last_col}{row_idx}",
                             f"  {si}. {sys['name']}：{note}",
                             bold=False, size=9, color=C_TEXT, bg=C_BLUE_LIGHT,
                             align="left", wrap=True)
                ws.row_dimensions[row_idx].height = 22
                row_idx += 1

    # ── 系统工作量明细表 ──
    row_idx += 1
    _merge_style(ws, f"A{row_idx}:{last_col}{row_idx}",
                 "各系统工作量及费用明细",
                 bold=True, size=11, color=C_WHITE, bg=C_BLUE_DARK, align="center")
    ws.row_dimensions[row_idx].height = 20
    row_idx += 1

    headers = [
        "序\n号", "系统名称", "级\n别", "模块\n数", "接口\n数", "数据\n表数",
        "数据量\n(GB)", "改造类型",
        "模块\n工时", "接口\n工时", "数据表\n工时", "改造\n附加量",
        "最终工\n作量(天)", "人工费\n(元)", "测试费\n(元)", "不含税\n(元)", "含税\n(元)"
    ]
    _apply_header_row(ws, row_idx, headers)
    ws.row_dimensions[row_idx].height = 42
    row_idx += 1

    sum_labor = sum_test = sum_excl = sum_incl = sum_work = 0.0
    for idx, sys in enumerate(d["systems"], 1):
        labor = sys["total_work"] * price_per_day
        test_c = labor * test_rate
        sub = labor + test_c
        mgmt = sub * management_rate
        risk = sub * risk_rate
        excl = sub + mgmt + risk
        incl = excl * (1 + tax_rate)
        sum_labor += labor; sum_test += test_c; sum_excl += excl; sum_incl += incl
        sum_work += sys["total_work"]

        adaptation_str = "、".join(sys.get("adaptation_type", [])) or "—"
        alt = (idx % 2 == 0)
        row = [
            idx, sys["name"], sys.get("level", "—"),
            sys.get("modules", 0), sys.get("interfaces", 0), sys.get("databases", 0),
            _num(sys.get("data_size_gb", 0)), adaptation_str,
            _num(sys.get("module_work", 0)), _num(sys.get("interface_work", 0)),
            _num(sys.get("db_work", 0)), _num(sys.get("adaptation_work", 0)),
            _num(sys["total_work"]),
            _num(labor), _num(test_c), _num(excl), _num(incl)
        ]
        _apply_data_row(ws, row_idx, row, alt=alt)
        row_idx += 1

    # 外部接口专项行
    ext_if = d.get("external_interfaces", 0)
    addon = d.get("interface_addon_work", 0)
    if ext_if > 0:
        ext_labor = addon * price_per_day
        ext_test = ext_labor * test_rate
        ext_sub = ext_labor + ext_test
        ext_excl = ext_sub * (1 + management_rate + risk_rate)
        ext_incl = ext_excl * (1 + tax_rate)
        sum_excl += ext_excl; sum_incl += ext_incl
        _apply_data_row(ws, row_idx,
                        ["*", f"外部对接接口专项（共{ext_if}个）", "—",
                         "—", ext_if, "—", "—", "接口改造",
                         "—", _num(addon), "—", "—",
                         _num(addon), _num(ext_labor), _num(ext_test), _num(ext_excl), _num(ext_incl)],
                        bold=False, color="CC3300")
        row_idx += 1

    # 小计行
    for ci in range(1, TOTAL_COLS + 1):
        c = ws.cell(row_idx, ci)
        c.font = _font(bold=True, size=10, color=C_WHITE)
        c.fill = _fill("1B5E20")
        c.alignment = _center()
        c.border = _header_border()
    ws.cell(row_idx, 2).value = "小  计"
    ws.cell(row_idx, 13).value = _num(sum_work)
    ws.cell(row_idx, 14).value = _num(sum_labor)
    ws.cell(row_idx, 15).value = _num(sum_test)
    ws.cell(row_idx, 16).value = _num(sum_excl)
    ws.cell(row_idx, 17).value = _num(sum_incl)
    row_idx += 1

    # ── 费用汇总与计算过程 ──
    row_idx += 1
    cost = d["cost"]
    _merge_style(ws, f"A{row_idx}:{last_col}{row_idx}",
                 "费用计算过程（含完整公式推导）",
                 bold=True, size=11, color=C_WHITE, bg=C_BLUE_DARK, align="center")
    ws.row_dimensions[row_idx].height = 20
    row_idx += 1

    tw = d["total_workload"]
    cost_steps = [
        ("①  总工作量",        f"{tw} 人天",                                 "各系统工作量之和（含外部接口专项）"),
        ("②  人工费",          f"¥{cost['labor_cost']:,.2f}",               f"{tw} 人天 × ¥{price_per_day:,}/天"),
        ("③  测试费",          f"¥{cost['test_cost']:,.2f}",                f"人工费 × {test_rate*100:.0f}%（含功能/兼容/回归测试）"),
        ("④  人工+测试 小计",  f"¥{cost['subtotal']:,.2f}",                 "② + ③"),
        ("⑤  项目管理费",      f"¥{cost['management_cost']:,.2f}",          f"小计 × {management_rate*100:.0f}%（项目经理、协调、文档）"),
        ("⑥  风险不可预见费",  f"¥{cost['risk_cost']:,.2f}",                f"小计 × {risk_rate*100:.0f}%（技术风险、需求变更储备）"),
        ("⑦  不含税合计",      f"¥{cost['total_excl_tax']:,.2f}",           "④ + ⑤ + ⑥"),
        ("⑧  增值税",          f"¥{cost['total_excl_tax']*(tax_rate):,.2f}", f"不含税合计 × {tax_rate*100:.0f}%（信息技术服务税率）"),
        ("⑨  含税合计（最终）",f"¥{cost['total_incl_tax']:,.2f}",           "⑦ × (1 + 税率)"),
    ]

    step_headers = ["费用项目", "金额", "计算说明"]
    _apply_header_row(ws, row_idx, step_headers, fills=[C_BLUE_DARK] * 3)
    ws.merge_cells(f"B{row_idx}:G{row_idx}")
    ws.merge_cells(f"H{row_idx}:{last_col}{row_idx}")
    row_idx += 1

    for ki, (step, amt, formula) in enumerate(cost_steps):
        is_final = "最终" in step
        is_subtotal = "小计" in step or "合计" in step
        bg = C_BLUE_DARK if is_final else (C_GRAY if is_subtotal else (C_BLUE_LIGHT if ki % 2 == 0 else C_WHITE))
        fc = C_WHITE if is_final else C_TEXT

        cell_s = ws.cell(row_idx, 1, step)
        cell_s.font = _font(bold=is_final or is_subtotal, size=10, color=fc)
        cell_s.fill = _fill(bg); cell_s.alignment = _left(); cell_s.border = _border()

        ws.merge_cells(f"B{row_idx}:G{row_idx}")
        cell_a = ws.cell(row_idx, 2, amt)
        cell_a.font = _font(bold=True, size=11, color=C_ORANGE if is_final else (C_GREEN if is_subtotal else C_TEXT))
        cell_a.fill = _fill(bg); cell_a.alignment = _center(); cell_a.border = _border()

        ws.merge_cells(f"H{row_idx}:{last_col}{row_idx}")
        cell_f = ws.cell(row_idx, 8, formula)
        cell_f.font = _font(size=9, color=fc if is_final else "555555")
        cell_f.fill = _fill(bg); cell_f.alignment = _left(wrap=True); cell_f.border = _border()
        ws.row_dimensions[row_idx].height = 18
        row_idx += 1

    # 大写金额
    _merge_style(ws, f"A{row_idx}:{last_col}{row_idx}",
                 f"含税合计（大写）：{_amount_cn(cost['total_incl_tax'])}",
                 bold=True, size=12, color=C_ORANGE, bg="FFF8E7", align="left")
    ws.row_dimensions[row_idx].height = 22
    row_idx += 1

    # ── 工时阶段分布 ──
    row_idx += 1
    _merge_style(ws, f"A{row_idx}:{last_col}{row_idx}",
                 "工时阶段分布（各系统）",
                 bold=True, size=11, color=C_WHITE, bg=C_BLUE_DARK, align="center")
    ws.row_dimensions[row_idx].height = 20
    row_idx += 1

    phases = list(PHASE_RATIOS.items())
    ph_headers = ["序号", "系统名称"] + [f"{p}\n({int(ratio*100)}%)" for p, ratio in phases] + ["合计(人天)"]
    _apply_header_row(ws, row_idx, ph_headers)
    ws.merge_cells(f"C{row_idx}:{get_column_letter(2+len(phases))}{row_idx}")
    row_idx += 1

    for si, sys in enumerate(d["systems"], 1):
        alt = si % 2 == 0
        ph_vals = [_num(sys.get("phases", {}).get(p, 0)) for p, _ in phases]
        _apply_data_row(ws, row_idx, [si, sys["name"]] + ph_vals + [_num(sys["total_work"])], alt=alt)
        row_idx += 1


def _sheet_params(wb):
    """评估参数说明 Sheet（完整参数体系+行业对标+调整建议）"""
    config = load_config()
    ws = wb.create_sheet("评估参数说明")
    ws.sheet_view.showGridLines = False

    col_widths = [4, 26, 14, 10, 42]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    TOTAL_COLS = len(col_widths)
    last_col = get_column_letter(TOTAL_COLS)

    _merge_style(ws, f"A1:{last_col}1", "评估参数说明（本次评估使用参数一览）",
                 bold=True, size=14, color=C_WHITE, bg=C_BLUE_DARK, align="center")
    ws.row_dimensions[1].height = 22

    # ── 基础工作量系数 ──
    _merge_style(ws, f"A2:{last_col}2", "一、基础工作量系数",
                 bold=True, size=11, color=C_WHITE, bg=C_BLUE, align="left")
    ws.row_dimensions[2].height = 18
    _apply_header_row(ws, 3, ["序号", "参数名（JSON Key）", "参数含义", "当前值", "说明与调整建议"])
    ws.row_dimensions[3].height = 18

    base_params = [
        ("module_factor",    "模块工作量系数",
         config.get("module_factor", 3.5),
         f"每个功能模块基础工作量（人天）。含功能开发、单元测试、文档。信创项目因改造难度建议3~5，当前值：{config.get('module_factor', 3.5)}"),
        ("interface_factor", "接口工作量系数",
         config.get("interface_factor", 1.5),
         "每个内部接口的改造工作量（人天）。含接口协议适配、联调、测试。建议1~2。"),
        ("table_factor",     "数据表工作量系数",
         config.get("table_factor", 0.8),
         "每张数据库表迁移/改造工作量（人天）。含表结构分析、SQL迁移、数据验证。建议0.5~1.5。"),
        ("data_factor",      "数据量系数",
         config.get("data_factor", 0.01),
         "每GB数据的迁移工作量（人天）。含抽取清洗、格式转换、比对验证。建议0.005~0.02。"),
    ]
    for ki, (key, name, val, desc) in enumerate(base_params, 1):
        alt = ki % 2 == 0
        _apply_data_row(ws, 3 + ki, [ki, key, name, val, desc], alt=alt)
    row_idx = 3 + len(base_params) + 1

    # ── 费率参数 ──
    _merge_style(ws, f"A{row_idx}:{last_col}{row_idx}", "二、费率参数",
                 bold=True, size=11, color=C_WHITE, bg=C_BLUE, align="left")
    ws.row_dimensions[row_idx].height = 18
    row_idx += 1
    _apply_header_row(ws, row_idx, ["序号", "参数名（JSON Key）", "参数含义", "当前值", "说明与行业参考"])
    ws.row_dimensions[row_idx].height = 18
    row_idx += 1

    rate_params = [
        ("price_per_day",    "人天综合费率",
         f"¥{config.get('price_per_day', 2200):,}",
         "每人天综合成本（元），含工资、社保、管理分摊、场地。信创人才稀缺，市场参考：高级工程师1800~2800元/天。"),
        ("test_rate",        "测试费率",
         f"{config.get('test_rate', 0.15)*100:.0f}%",
         "测试费 = 人工费 × 测试费率。含功能测试、兼容性测试（信创环境）、回归测试。行业参考：10%~20%。"),
        ("management_rate",  "项目管理费率",
         f"{config.get('management_rate', 0.12)*100:.0f}%",
         "项目管理费 = (人工+测试小计) × 管理费率。含项目经理、计划管控、风险管理、进度汇报。行业参考：8%~15%。"),
        ("risk_rate",        "风险不可预见费率",
         f"{config.get('risk_rate', 0.10)*100:.0f}%",
         "风险费 = (人工+测试小计) × 风险费率。覆盖技术不确定性、需求变更、第三方依赖风险。行业参考：5%~15%。"),
        ("tax_rate",         "增值税率",
         f"{config.get('tax_rate', 0.09)*100:.0f}%",
         "信息技术服务适用税率9%（一般纳税人）。小规模纳税人可按3%或申请免税，请结合合同实际情况调整。"),
    ]
    for ki, (key, name, val, desc) in enumerate(rate_params, 1):
        alt = ki % 2 == 0
        _apply_data_row(ws, row_idx, [ki, key, name, val, desc], alt=alt)
        row_idx += 1
    row_idx += 1

    # ── 改造类型附加量 ──
    _merge_style(ws, f"A{row_idx}:{last_col}{row_idx}", "三、改造类型附加工作量参数（adaptation_factors）",
                 bold=True, size=11, color=C_WHITE, bg=C_BLUE, align="left")
    ws.row_dimensions[row_idx].height = 18
    row_idx += 1
    _apply_header_row(ws, row_idx, ["序号", "改造类型", "适用场景", "当前值(人天)", "调整建议"])
    ws.row_dimensions[row_idx].height = 18
    row_idx += 1

    adapt_guidance = {
        "数据库替换":   ("将Oracle/MySQL等替换为达梦、人大金仓、GaussDB等国产数据库", "系统越复杂存储过程越多，建议50~120人天"),
        "ETL平台适配":  ("Kettle/DataStage等ETL工具替换为国产FineDataLink/DataWorks等", "工作流越多建议适当提高，参考40~80人天"),
        "数据迁移":     ("历史业务数据抽取、清洗、格式转换、完整性校验与迁移", "数据量及质量问题影响大，建议20~60人天"),
        "硬件架构适配": ("x86应用向飞腾/鲲鹏(ARM)架构移植，含编译适配与性能测试", "C/C++代码越多工作量越高，建议15~40人天"),
        "操作系统适配": ("在麒麟/统信UOS等国产OS上部署调试，含依赖库替换", "依赖第三方库越少越简单，建议10~30人天"),
        "安全加固":     ("国密SM2/SM4算法集成、接口鉴权、日志审计、等保整改", "等保级别越高工作量越大，建议10~25人天"),
        "中间件替换":   ("RabbitMQ/Redis等替换为国产消息队列/缓存组件", "消息量和依赖程度影响，建议8~20人天"),
        "前端框架适配": ("前端UI兼容信创浏览器，含字体、打印、控件适配", "Vue/React框架下相对简单，建议8~18人天"),
        "接口改造":     ("第三方/外部接口协议适配、认证改造、联调测试", "按系统计，外部接口数量单独计算附加量"),
    }
    adapt_factors = config.get("adaptation_factors", {})
    for ki, (k, v) in enumerate(adapt_factors.items(), 1):
        alt = ki % 2 == 0
        scene, guidance = adapt_guidance.get(k, ("—", "—"))
        _apply_data_row(ws, row_idx, [ki, k, scene, v, guidance], alt=alt)
        ws.row_dimensions[row_idx].height = 18
        row_idx += 1
    row_idx += 1

    # ── 系数说明 ──
    _merge_style(ws, f"A{row_idx}:{last_col}{row_idx}", "四、级别系数与复杂度系数说明",
                 bold=True, size=11, color=C_WHITE, bg=C_BLUE, align="left")
    ws.row_dimensions[row_idx].height = 18
    row_idx += 1
    level_notes = [
        ("级别系数-核心", "1.3", "直接影响业务连续性的核心生产/交易系统，改造风险最高，需重点保障"),
        ("级别系数-重要", "1.1", "关键业务支撑系统，停机影响较大，需要专项测试保障"),
        ("级别系数-一般", "1.0", "辅助/管理类系统，停机影响可控，标准改造流程"),
        ("复杂度加成-模块数>20", "+0.15", "功能模块多说明系统业务逻辑复杂，改造测试覆盖面广"),
        ("复杂度加成-接口数>30", "+0.10", "接口多说明系统与外部耦合度高，改造风险和联调成本高"),
        ("复杂度加成-接口数>10", "+0.05", "中等耦合度，有一定联调工作量"),
        ("复杂度加成-改造类型≥4", "+0.20", "同时涉及多类改造，交叉影响大，协调成本高"),
        ("复杂度加成-改造类型≥2", "+0.10", "涉及两种以上改造，有一定综合影响"),
    ]
    _apply_header_row(ws, row_idx, ["序号", "系数项目", "系数值", "说明"])
    ws.merge_cells(f"C{row_idx}:D{row_idx}")
    ws.merge_cells(f"E{row_idx}:{last_col}{row_idx}")
    ws.row_dimensions[row_idx].height = 18
    row_idx += 1
    for ki, (name, val, desc) in enumerate(level_notes, 1):
        alt = ki % 2 == 0
        bg = C_BLUE_LIGHT if alt else C_WHITE
        c1 = ws.cell(row_idx, 1, ki); c1.font=_font(size=9); c1.fill=_fill(bg); c1.alignment=_center(); c1.border=_border()
        c2 = ws.cell(row_idx, 2, name); c2.font=_font(bold=True, size=9, color=C_BLUE_DARK); c2.fill=_fill(bg); c2.alignment=_left(); c2.border=_border()
        ws.merge_cells(f"C{row_idx}:D{row_idx}")
        c3 = ws.cell(row_idx, 3, val); c3.font=_font(bold=True, size=10, color=C_ORANGE); c3.fill=_fill(bg); c3.alignment=_center(); c3.border=_border()
        ws.merge_cells(f"E{row_idx}:{last_col}{row_idx}")
        c5 = ws.cell(row_idx, 5, desc); c5.font=_font(size=9, color=C_TEXT); c5.fill=_fill(bg); c5.alignment=_left(wrap=True); c5.border=_border()
        ws.row_dimensions[row_idx].height = 18
        row_idx += 1


def _amount_cn(amount):
    """金额转中文大写（简化版）"""
    units = ["", "万", "亿"]
    amount = round(amount, 2)
    if amount >= 1_0000_0000:
        return f"人民币{amount/1_0000_0000:.4f}亿元整"
    elif amount >= 1_0000:
        wan = amount / 10000
        return f"人民币{wan:.2f}万元整"
    else:
        return f"人民币{amount:.2f}元整"


if __name__ == "__main__":
    app.run(debug=True)
